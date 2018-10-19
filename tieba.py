#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import re
import json
import time
import pandas
import requests
from datetime import datetime
from bs4 import BeautifulSoup


class Baidu(object):

    def __init__(self, url):
        self.tid = int(re.search(r'tieba.baidu.com/p/(\d+)', url)[1]) # 帖子id
        self.url = 'http://tieba.baidu.com/p/'+str(self.tid)+'?pn={}' # 帖子链接，可翻页
        self.date = datetime(1, 1, 1)
        self.flag = True # 是否404，默认有内容
        self.count = 0 # 总楼层计数
        self.t_count = 0 #内部人员计数
        self.url_x = 'https://tieba.baidu.com/p/comment?tid='+str(self.tid)+'&pid={}&pn={}' # 楼中楼链接
        self.t_list = [] # 内部人员id列表，隐去
        self.get_info() # 进行判断

    # 内部人员回帖计数函数
    def t_ct(self, id):
        if id in self.t_list:
            self.t_count += 1

    # 检查是否404，有内容就获取tie_n, ye_n, title
    def get_info(self):
        soup = self.get_soup(self.url.format(1))
        if soup.select('.page404'):
            self.flag = False # 该贴删除
        else:
            self.tie_n = int(soup.select('.l_reply_num .red')[0].text)
            self.ye_n = int(soup.select('.l_reply_num .red')[1].text)
            self.title = soup.select('.core_title_txt')[0].get('title')

    # 最新回帖时间    
    def c_time(self, t1, t2):
        t2 = datetime.strptime(t2, '%Y-%m-%d %H:%M')
        return t2 if t1 < t2 else t1

    # 获取页面
    def get_soup(self, url):
        res = requests.get(url)
        soup = BeautifulSoup(res.text, 'html.parser')
        return soup

    # 获取用户id
    def get_user_id(self, home_page):
        try:
            hp = 'http://tieba.baidu.com{}'
            s = BeautifulSoup(requests.get(hp.format(home_page)).text,'lxml')
            # time.sleep(1)
            return int(s.select('.mygift-more')[0].get('data-user-id'))
        except:
            print('个人主页404')
            return 0
    
    # 获取楼中楼内容
    def get_content_in_floor(self, n, pid, cn):
        r = []
        rank = 0
        for p in range(1, int(cn/10)+2):
            s = BeautifulSoup(requests.get(self.url_x.format(pid,p)).text,'lxml')
            # time.sleep(1)
            for i in s.select('.lzl_single_post'):
                d = {'帖子id':self.tid,'帖子标题':self.title}
                rank += 1
                d['楼层'] = float(str(n)+'.'+str(rank))
                d['昵称'] = i.select('.at.j_user_card')[0].get('username')
                d['昵称id'] = self.get_user_id(i.select('.at.j_user_card')[0].get('href'))
                d['回帖内容'] = i.select('.lzl_content_main')[0].text.strip()
                d['回帖时间'] = i.select('.lzl_time')[0].text
                self.date = self.c_time(self.date, d['回帖时间'])
                self.t_ct(d['昵称id'])
                r.append(d)
        return r

    # 获取楼层内容
    def get_floor(self, floor):
        r = []
        j = json.loads(floor.get('data-field'))
        d = {'帖子id':self.tid,'帖子标题':self.title}
        d['楼层'] = j['content']['post_no']
        d['昵称id'] = j['author']['user_id']
        d['昵称'] = j['author']['user_name']
        d['回帖内容'] = floor.select('.d_post_content')[0].text.strip()
        try:
            d['回帖时间'] = j['content']['date']
        except:
            d['回帖时间'] =floor.select('span.tail-info')[-1].text
        r.append(d)
        self.date = self.c_time(self.date, d['回帖时间'])
        if d['楼层']>1:
            self.t_ct(d['昵称id'])
        cn = j['content']['comment_num']
        # 楼中楼判断
        if cn > 0:
            pid = j['content']['post_id']
            r.extend(self.get_content_in_floor(d['楼层'],pid,cn))
        return r

    # 获取帖子内其他人员总数
    def get_other_num(self, r_list):
        df = pandas.DataFrame(r_list)
        self.s_num = len(set(df['昵称id'])-set(self.t_list))

    # 循环获取整个帖子内容，包括翻页
    def get_content(self):
        if self.flag:
            r = []
            for p in range(1, self.ye_n+1):
                s = self.get_soup(self.url.format(p))
                # time.sleep(5)
                t_all = s.select('.l_post.j_l_post.l_post_bright')
                self.count += len(t_all)
                for i in t_all:
                    r.extend(self.get_floor(i))
            self.get_other_num(r)
            return self.count, self.tie_n, self.t_count, self.tie_n-self.t_count, self.s_num, self.date.strftime('%Y-%m-%d %H:%M'), r
        else:
            print('很抱歉，该贴已被删除。',self.url.replace('?pn={}',''))
            return None, None, None, None, None, None, None

# 总楼层，总回帖，内部回帖，其他回帖，最新时间，所有信息r
# r：帖子id，帖子标题，回帖id，楼层，回帖昵称，回帖内容，回帖时间

# 获取帖子链接，底表链接在第3列
df = pandas.read_excel(r'C:\Users\Administrator\Desktop\tieba\底表.xlsx')
dt = pandas.DataFrame()
df['总楼层'], df['总回帖'], df['内部回帖'], df['其他回帖'], df['其他人数'], df['最新回帖时间'] = None, None, None, None, None, None

for i in range(len(df)):
    try:
        c = Baidu(df.iloc[i,2])
        df.iloc[i,3],df.iloc[i,4],df.iloc[i,5],df.iloc[i,6],df.iloc[i,7], df.iloc[i,8], r = c.get_content()
        dt = dt.append(r)
        if i%10 == 0:
            percent = (i+1)/len(df)*100
            print('已爬取：%6.2f%%' % percent)
    except Exception as e:
        print(e)
        print('错误行',i+2)

print('已爬取：%6.2f%%' % 100)

# 已爬取时间为准，获取贴子在贴吧的排名，只看前5页，返回前250条链接的排名字典
def rank_dict():
    url = 'http://tieba.baidu.com/f?ie=utf-8&kw=%E5%B0%9A%E5%BE%B7%E6%9C%BA%E6%9E%84&pn={}'
    rank = 0
    d = {}
    for p in range(5):
        tl = BeautifulSoup(requests.get(url.format(p*50)).text, 'html.parser').select('.j_thread_list')
        for i in tl:
            rank += 1
            j = json.loads(i.get('data-field'))
            d[j['id']] = rank
    return d

# 得到爬取帖子的排名
def get_rank(url, r_dict):
    tid = int(re.search(r'tieba.baidu.com/p/(\d+)', url)[1])
    if tid in r_dict.keys():
        return r_dict[tid]
    else:
        return None

rd = rank_dict()
df['排名'] = df['对应链接'].apply(lambda x: get_rank(x, rd))
df.sort_values(by=['排名','其他人数'],inplace=True)
df.fillna('-',inplace=True)
t_list = [1630857641,1346648388,3410935512,918846377,1681283402,948515849,528410350,1717490373,1653928153,3157505144,447421476,1326252398,202224677,991945491,980887861,816219351,3838890791,2795053047,2782478011,3447753151,3838920269,703342274,3106538569,3845815164,1264209080,1974099029,22587715,3844621726,2976057994,1908166708,208101694,1228344732,2670663946]
dt = dt[['帖子id', '帖子标题', '楼层', '昵称id', '昵称', '回帖内容', '回帖时间']]
total_other_num = len(set(dt['昵称id'])-set(t_list))
today = datetime.today().strftime('%m%d')
'''
writer = pandas.ExcelWriter('C:/Users/Administrator/Desktop/tieba/result/百度贴吧结果'+today+'.xlsx')
df.to_excel(writer,'成表',index=0)
dt.to_excel(writer,'详情',index=0)
writer.save()
'''

# 以下都是写入excel，设定特定格式，不详细说了
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

wb = Workbook()
print('写入数据中...')

ws = wb.active
ws.title = "成表"
'''
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

ws.cell(row=1,column=12).value = '总其他人数'
ws.cell(row=2,column=12).value = total_other_num
'''
n = 2
for r in dataframe_to_rows(df, index=False, header=True):
    for m in range(2,12):
        ws.cell(row=n,column=m).value = r[m-2]
        if n == 2:
            ws.cell(row=n, column=m).font = Font(bold=True, color='FFFFFFFF')
            ws.cell(row=n, column=m).fill = PatternFill(fill_type='solid', start_color='FF6176AA')
        elif n == len(df)+2:
            ws.cell(row=n, column=m).border = Border(bottom=Side(border_style='thin', color='FF000000'))
        else:
            ws.cell(row=n, column=m).border = Border(bottom=Side(border_style='thin', color='FFD9D9D9'))
        if m in [5,6,7,8,9,10,11]:
            ws.cell(row=n, column=m).alignment = Alignment(horizontal='right')
    n = n+1
ws.cell(row=2,column=10).alignment = Alignment(horizontal='left')
ws.cell(row=2,column=13).value = '总其他人数'
ws.cell(row=2,column=13).font = Font(bold=True, color='FFFFFFFF')
ws.cell(row=2,column=13).fill = PatternFill(fill_type='solid', start_color='FF6176aa', end_color='FF000000')
ws.cell(row=3,column=13).value = total_other_num
ws.cell(row=3,column=13).border = Border(bottom=Side(border_style='thin', color='FF000000'))

ws.column_dimensions['B'].width = 9
ws.column_dimensions['C'].width = 9
ws.column_dimensions['D'].width = 34
ws.column_dimensions['E'].width = 9
ws.column_dimensions['F'].width = 9
ws.column_dimensions['G'].width = 9
ws.column_dimensions['H'].width = 9
ws.column_dimensions['I'].width = 9
ws.column_dimensions['J'].width = 15.14
ws.column_dimensions['K'].width = 6
ws.column_dimensions['M'].width = 12

ws.sheet_view.showGridLines = False #不显示网格线


ws2 = wb.create_sheet(title="详情")
for r in dataframe_to_rows(dt, index=False, header=True):
    ws2.append(r)

name = 'C:/Users/Administrator/Desktop/tieba/result/百度贴吧结果'+today+'.xlsx'
wb.save(name)
print('写入成功！')
